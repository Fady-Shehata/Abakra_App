# -*- coding: utf-8 -*-
"""
Build the faithful all-200 Excel workbook for the church-councils question bank.

Source of question text + per-slot option order:
  * Q1-50   and Q186-200 : questions_data.py  (genuine, distinct slots)
  * Q51-185             : pdata/pNN.py         (per-page verified transcriptions)

Every numbered slot is reproduced exactly as it appears in the PDF (repeats
included), each with its own option order.  Answers and difficulty levels are
LOCKED (cross-checked against the official key images, pages 35-37).
"""
import os
import re
import glob
import importlib.util
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ---- 1. Load genuine slots (Q1-50, Q186-200) from questions_data.py --------
spec = importlib.util.spec_from_file_location("questions_data", os.path.join(BASE, "questions_data.py"))
qd = importlib.util.module_from_spec(spec)
spec.loader.exec_module(qd)
Q = qd.Q

# ---- 2. Load verified per-slot transcriptions (Q51-185) from pdata/ ---------
pdata = {}
for path in sorted(glob.glob(os.path.join(BASE, "pdata", "p*.py"))):
    s = importlib.util.spec_from_file_location(os.path.basename(path)[:-3], path)
    m = importlib.util.module_from_spec(s)
    s.loader.exec_module(m)
    for n, v in m.D.items():
        if n in pdata:
            raise RuntimeError(f"Duplicate slot {n} across pdata files ({path})")
        pdata[n] = v

# ---- 3. Assemble the final 200 slots ---------------------------------------
QUESTIONS = {}
for n in range(1, 51):
    QUESTIONS[n] = Q[n]
for n in range(51, 186):
    QUESTIONS[n] = pdata[n]
for n in range(186, 201):
    QUESTIONS[n] = Q[n]

missing = [n for n in range(1, 201) if n not in QUESTIONS]
assert not missing, f"Missing slots: {missing}"

# ---- 4. LOCKED answers (q:letter) ------------------------------------------
ANS_STR = (
 "1أ 2أ 3د 4ج 5ب 6ج 7ب 8ج 9ب 10ج 11ج 12أ 13د 14ب 15د 16أ 17د 18أ 19ب 20د "
 "21ج 22د 23ب 24أ 25ج 26ج 27د 28ج 29ج 30ج 31د 32أ 33ب 34أ 35ب 36ب 37أ 38د 39د 40ج "
 "41ج 42ب 43ج 44ب 45د 46ج 47أ 48أ 49ب 50ب 51أ 52ج 53أ 54أ 55د 56د 57ب 58ج 59ب 60ب "
 "61أ 62ب 63ج 64أ 65أ 66أ 67أ 68ج 69أ 70أ 71ب 72ج 73ج 74ب 75ب 76ب 77ب 78ب 79د 80ب "
 "81ب 82ب 83د 84ب 85أ 86د 87د 88ب 89ب 90د 91ج 92ج 93ب 94ج 95ب 96أ 97ج 98د 99د 100د "
 "101أ 102أ 103د 104أ 105ب 106د 107أ 108ب 109ب 110أ 111ب 112أ 113أ 114ج 115ب 116ب 117ب 118ب 119ب 120ب "
 "121د 122د 123ب 124ب 125ج 126ب 127ج 128أ 129ب 130أ 131د 132أ 133ب 134د 135د 136د 137أ 138د 139ج 140أ "
 "141ج 142أ 143د 144د 145ب 146ج 147ب 148ج 149ب 150ج 151ب 152ب 153أ 154أ 155ج 156د 157ب 158أ 159ج 160ب "
 "161د 162ب 163د 164ب 165د 166د 167ج 168ج 169د 170ب 171د 172ب 173ب 174د 175ج 176د 177ب 178ج 179د 180د "
 "181ب 182د 183أ 184ج 185أ 186ب 187أ 188ج 189ج 190ب 191ج 192أ 193د 194ب 195د 196ج 197ب 198ب 199ج 200ب"
)
ANSWERS = {}
for tok in ANS_STR.split():
    mm = re.match(r"(\d+)(.+)", tok)
    ANSWERS[int(mm.group(1))] = mm.group(2)
assert len(ANSWERS) == 200, f"Expected 200 answers, got {len(ANSWERS)}"

# ---- 5. LOCKED difficulty levels (المستوى) ---------------------------------
LEVELS = {}
lv1_50 = [
 "متوسط", "سهل", "متوسط", "متوسط", "متوسط", "سهل", "صعب", "سهل", "سهل", "صعب",
 "صعب", "متوسط", "صعب", "صعب", "متوسط", "متوسط", "صعب", "متوسط", "سهل", "متوسط",
 "سهل", "صعب", "متوسط", "سهل",
 "سهل", "سهل", "صعب", "متوسط", "سهل", "سهل", "صعب", "سهل", "سهل", "سهل",
 "متوسط", "صعب", "متوسط", "صعب", "متوسط", "صعب", "سهل", "متوسط", "صعب", "سهل",
 "سهل", "سهل", "سهل", "سهل", "متوسط", "سهل",
]
assert len(lv1_50) == 50
for i, v in enumerate(lv1_50, start=1):
    LEVELS[i] = v
for n in range(51, 109):
    LEVELS[n] = "سهل"
for n in range(109, 162):
    LEVELS[n] = "متوسط"
for n in range(162, 201):
    LEVELS[n] = "صعب"
assert len(LEVELS) == 200

# ---- 6. Build rows + validate ----------------------------------------------
IDX = {"أ": 0, "ب": 1, "ج": 2, "د": 3}
rows = []
for n in range(1, 201):
    stem, opts = QUESTIONS[n]
    assert len(opts) == 4 and all(o.strip() for o in opts), f"Bad options at slot {n}"
    ans = ANSWERS[n]
    assert ans in IDX, f"Bad answer letter '{ans}' at slot {n}"
    assert opts[IDX[ans]].strip(), f"Answer {ans} points to empty option at slot {n}"
    rows.append((n, stem, opts[0], opts[1], opts[2], opts[3], ans, LEVELS[n]))
assert len(rows) == 200

# ---- 7. Write the styled RTL workbook --------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "بنك الأسئلة"
ws.sheet_view.rightToLeft = True

headers = ["رقم السؤال", "السؤال", "أ", "ب", "ج", "د", "الإجابة الصحيحة", "المستوى"]
ws.append(headers)

hf = PatternFill("solid", fgColor="1F3864")
hfont = Font(bold=True, color="FFFFFF", size=12)
thin = Side(style="thin", color="B0B0B0")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
rgt = Alignment(horizontal="right", vertical="center", wrap_text=True)

for c in ws[1]:
    c.fill = hf
    c.font = hfont
    c.alignment = ctr
    c.border = bd

for r in rows:
    ws.append(list(r))

for col, w in {"A": 10, "B": 58, "C": 24, "D": 24, "E": 24, "F": 24, "G": 16, "H": 12}.items():
    ws.column_dimensions[col].width = w

for r in range(2, 202):
    for col in range(1, 9):
        cell = ws.cell(row=r, column=col)
        cell.border = bd
        cell.alignment = rgt if col == 2 else ctr
    ws.row_dimensions[r].height = 30

ws.freeze_panes = "A2"
out = os.path.join(BASE, "output", "بنك_الأسئلة_المجامع_الكنسية_v3.xlsx")
try:
    wb.save(out)
except PermissionError:
    raise SystemExit(f"Cannot write '{out}' — please close it in Excel and re-run.")
print("Saved", out)
print("Total rows (incl. header):", ws.max_row)
