# -*- coding: utf-8 -*-
import os
import re

import fitz
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PDF_PATH = os.path.join(BASE, "data", "رسالة العبرانين1.pdf")
OUT_PATH = os.path.join(BASE, "output", "رسالة_العبرانين1_اختر_الاجابة_الصحيحة.xlsx")

OPTION_RE = re.compile(r"^([أبجد])\)\s*(.+)$")
QUESTION_NUM_RE = re.compile(r"^(\d+)\s*[\.-]\s*(.+)$")
OPTION_MARKER_RE = re.compile(r"([أبجد])\)\s*")


def clean_text(s):
    s = s.strip()
    s = s.replace("**", "")
    s = s.replace("*", "")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def intersects(a, b):
    x0 = max(a.x0, b.x0)
    y0 = max(a.y0, b.y0)
    x1 = min(a.x1, b.x1)
    y1 = min(a.y1, b.y1)
    return x1 > x0 and y1 > y0


def is_yellow_fill(fill):
    if not fill or not isinstance(fill, tuple) or len(fill) < 3:
        return False
    r, g, b = fill[:3]
    return r > 0.8 and g > 0.8 and b < 0.45


def extract_lines(page):
    yellow_rects = []
    for d in page.get_drawings():
        if is_yellow_fill(d.get("fill")) and d.get("rect") is not None:
            yellow_rects.append(d["rect"])

    lines = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for ln in block.get("lines", []):
            spans = ln.get("spans", [])
            txt = "".join(sp.get("text", "") for sp in spans).strip()
            if not txt:
                continue
            bbox = fitz.Rect(ln["bbox"])
            txt = clean_text(txt)
            if not txt:
                continue
            highlighted = any(intersects(bbox, r) for r in yellow_rects)
            lines.append((bbox.y0, txt, highlighted))

    lines.sort(key=lambda x: x[0])
    return lines


def finalize_question(current, questions):
    if not current:
        return
    if not current.get("question"):
        return
    opts = current.get("options", {})
    if not opts:
        return

    qtext = current["question"]
    answer_letter = current.get("answer_letter", "")
    if not answer_letter and current.get("answer_text"):
        # Fallback: infer by exact option text match if highlight was missed.
        for letter, text in opts.items():
            if text.strip() == current["answer_text"].strip():
                answer_letter = letter
                break

    questions.append(
        {
            "number": current.get("number"),
            "question": qtext,
            "A": opts.get("أ", ""),
            "B": opts.get("ب", ""),
            "C": opts.get("ج", ""),
            "D": opts.get("د", ""),
            "answer": answer_letter,
        }
    )


def extract_option(text):
    m = OPTION_RE.match(text)
    if not m:
        return None
    return m.group(1), m.group(2).strip()


def extract_multi_options(text):
    matches = list(OPTION_MARKER_RE.finditer(text))
    if not matches:
        return []

    out = []
    for i, m in enumerate(matches):
        letter = m.group(1)
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        val = text[start:end].strip()
        if val:
            out.append((letter, val))
    return out


def is_answer_marker(text):
    return "اإلجابة الصحيحة" in text or "الإجابة الصحيحة" in text


def is_header_line(text):
    return "أسئلة رسالة العبرانيين" in text or "المجموعة" in text


def normalize_question_text(text):
    text = text.strip()
    m = QUESTION_NUM_RE.match(text)
    if m:
        text = m.group(2).strip()
    text = re.sub(r"^[\.\-\)\(\s]+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def norm_match(s):
    s = re.sub(r"\s+", " ", str(s)).strip().lower()
    s = re.sub(r"[\.,:;\-\"'“”‘’()\[\]{}]+", "", s)
    return s


def parse_questions(pdf_path):
    doc = fitz.open(pdf_path)
    questions = []

    current = None
    running_num = 0
    mode = "seek_question"
    question_buffer = []
    last_option_letter = None

    for page in doc:
        lines = extract_lines(page)
        for _, raw_text, highlighted in lines:
            text = clean_text(raw_text)
            if not text or is_header_line(text):
                continue

            opt = extract_option(text)

            if mode == "seek_question":
                multi = extract_multi_options(text)
                if multi:
                    finalize_question(current, questions)

                    running_num += 1
                    qtext = normalize_question_text(" ".join(question_buffer))
                    question_buffer = []

                    current = {
                        "number": running_num,
                        "question": qtext,
                        "options": {},
                        "answer_letter": "",
                        "answer_text": "",
                    }
                    for letter, val in multi:
                        current["options"][letter] = val
                        last_option_letter = letter
                    mode = "collect_options"
                else:
                    if not is_answer_marker(text):
                        question_buffer.append(text)
                continue

            if mode == "collect_options":
                if is_answer_marker(text):
                    mode = "seek_answer"
                    last_option_letter = None
                    continue

                multi = extract_multi_options(text)
                if multi:
                    for letter, val in multi:
                        current["options"][letter] = val
                        last_option_letter = letter
                else:
                    # Continuation line for long option text.
                    if last_option_letter and last_option_letter in current["options"]:
                        current["options"][last_option_letter] = (
                            current["options"][last_option_letter] + " " + text
                        ).strip()
                continue

            if mode == "seek_answer":
                if opt:
                    current["answer_text"] = opt[1]
                    if highlighted:
                        current["answer_letter"] = opt[0]
                    elif not current.get("answer_letter"):
                        current["answer_letter"] = opt[0]

                    finalize_question(current, questions)
                    current = None
                    mode = "seek_question"
                    question_buffer = []
                    last_option_letter = None
                else:
                    # If OCR drops the option marker, try matching highlighted text to options.
                    if highlighted and current is not None:
                        current["answer_text"] = text
                        nt = norm_match(text)
                        for letter, opt_text in current["options"].items():
                            no = norm_match(opt_text)
                            if nt and (nt == no or nt in no or no in nt):
                                current["answer_letter"] = letter
                                break
                        finalize_question(current, questions)
                        current = None
                        mode = "seek_question"
                        question_buffer = []
                        last_option_letter = None
                    else:
                        # If OCR splits answer onto another line, keep looking.
                        continue

    finalize_question(current, questions)
    return questions


def write_excel(questions, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "اختر الإجابة الصحيحة"
    ws.sheet_view.rightToLeft = True

    headers = ["رقم السؤال", "السؤال", "أ", "ب", "ج", "د", "الإجابة الصحيحة", "المستوى"]
    ws.append(headers)

    hf = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="B0B0B0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for c in ws[1]:
        c.fill = hf
        c.font = hfont
        c.alignment = ctr
        c.border = bd

    for idx, q in enumerate(questions, start=1):
        ws.append([idx, q["question"], q["A"], q["B"], q["C"], q["D"], q["answer"], "اختر الإجابة الصحيحة"])

    for col, width in {"A": 10, "B": 58, "C": 24, "D": 24, "E": 24, "F": 24, "G": 16, "H": 20}.items():
        ws.column_dimensions[col].width = width

    for r in range(2, ws.max_row + 1):
        for col in range(1, 9):
            cell = ws.cell(row=r, column=col)
            cell.border = bd
            cell.alignment = rgt if col == 2 else ctr
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A2"

    try:
        wb.save(out_path)
    except PermissionError:
        raise SystemExit(f"Cannot write '{out_path}' - please close it in Excel and re-run.")


if __name__ == "__main__":
    questions = parse_questions(PDF_PATH)
    if not questions:
        raise SystemExit("No questions parsed from PDF.")

    write_excel(questions, OUT_PATH)

    answered = sum(1 for q in questions if q.get("answer"))
    print("Saved", OUT_PATH)
    print("Parsed questions:", len(questions))
    print("Detected answers:", answered)
