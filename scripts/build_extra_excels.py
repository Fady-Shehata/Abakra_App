# -*- coding: utf-8 -*-
"""
Build two extra Excel workbooks that share the exact structure/style of
  output/final/رسالة_العبرانين1_اختر_الاجابة_الصحيحة.xlsx

  1) معلومات عامة  (open question/answer bank)  -> data/معلومات عامة.txt
  2) قدرات ذهنية   (multiple-choice bank)        -> data/قدرات ذهنية.txt

Columns (RTL): رقم السؤال | السؤال | أ | ب | ج | د | الإجابة الصحيحة | المستوى
"""
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")
OUT = os.path.join(BASE, "output", "final")

AR_LETTERS = "أبجد"


def parse_general(path):
    """Parse the open Q&A file (three mixed layouts) into (question, answer) pairs."""
    rows = []
    pending_q = None
    with open(path, encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            if re.fullmatch(r"\d+\.?", line):          # a bare number marker
                continue
            if line.startswith("الإجابة:") or line.startswith("ج:"):
                ans = line.split(":", 1)[1].strip()
                if pending_q:
                    rows.append((pending_q, ans))
                    pending_q = None
                continue
            # otherwise it is a question line
            q = line
            if q.startswith("س:"):
                q = q.split(":", 1)[1].strip()
            pending_q = q
    return rows


def parse_mcq(path):
    """Parse the multiple-choice file into (stem, [a,b,c,d], answer_letter)."""
    with open(path, encoding="utf-8") as f:
        lines = [l.rstrip("\n") for l in f]

    # split into blocks that each start with a "N)" line
    blocks, cur = [], None
    for line in lines:
        if re.match(r"^\s*\d+\)", line):
            if cur is not None:
                blocks.append(cur)
            cur = [line]
        elif cur is not None:
            cur.append(line)
    if cur is not None:
        blocks.append(cur)

    rows = []
    for block in blocks:
        opts = {}
        answer = None
        answer_text = None
        stem_lines = []
        first_opt_seen = False
        for i, line in enumerate(block):
            s = line.strip()
            m_opt = re.match(r"^([أبجد])\)\s*(.*)$", s)
            m_ans = ("الإجابة" in s)
            if i == 0:
                s = re.sub(r"^\s*\d+\)\s*", "", s)  # strip the "N)" prefix
            if m_ans:
                after = s.split("الإجابة", 1)[1].lstrip(" :")
                mm = re.search(r"[أبجد]", after)
                if mm:
                    answer = mm.group(0)
                # text portion after the "letter)" for typo-tolerant matching
                mt = re.search(r"\)\s*(.+)$", after)
                if mt:
                    answer_text = mt.group(1).strip()
                continue
            if m_opt:
                first_opt_seen = True
                opts[m_opt.group(1)] = m_opt.group(2).strip()
                continue
            if not first_opt_seen and s:
                stem_lines.append(s)
        # typo tolerance: if the answer letter is missing/invalid, match by value
        if (answer is None or answer not in AR_LETTERS) and answer_text:
            for letter, val in opts.items():
                if val.strip() == answer_text:
                    answer = letter
                    break
        if len(opts) == 4 and answer is not None and answer in AR_LETTERS:
            stem = " ".join(stem_lines).strip()
            rows.append((stem, [opts["أ"], opts["ب"], opts["ج"], opts["د"]], answer))
    return rows


def write_workbook(path, header_rows, level_label):
    """header_rows: list of (num, stem, a, b, c, d, answer, level)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = level_label
    ws.sheet_view.rightToLeft = True

    headers = ["رقم السؤال", "السؤال", "أ", "ب", "ج", "د", "الإجابة الصحيحة", "المستوى"]
    ws.append(headers)

    hf = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="B0B0B0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for c in ws[1]:
        c.fill = hf
        c.font = hfont
        c.alignment = ctr
        c.border = bd

    for r in header_rows:
        ws.append(list(r))

    for col, w in {"A": 10, "B": 58, "C": 24, "D": 24, "E": 24, "F": 24, "G": 16, "H": 20}.items():
        ws.column_dimensions[col].width = w

    for r in range(2, len(header_rows) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=r, column=col)
            cell.border = bd
            cell.alignment = rgt if col == 2 else ctr
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Saved {path}  ({len(header_rows)} questions)")


def main():
    os.makedirs(OUT, exist_ok=True)

    # 1) معلومات عامة  (open Q&A – no options)
    general = parse_general(os.path.join(DATA, "معلومات عامة.txt"))
    g_rows = [
        (i, q, "", "", "", "", ans, "معلومات عامة")
        for i, (q, ans) in enumerate(general, start=1)
    ]
    write_workbook(os.path.join(OUT, "معلومات_عامة.xlsx"), g_rows, "معلومات عامة")

    # 2) قدرات ذهنية  (multiple choice)
    mcq = parse_mcq(os.path.join(DATA, "قدرات ذهنية.txt"))
    m_rows = [
        (i, stem, o[0], o[1], o[2], o[3], ans, "قدرات ذهنية")
        for i, (stem, o, ans) in enumerate(mcq, start=1)
    ]
    write_workbook(os.path.join(OUT, "قدرات_ذهنية.xlsx"), m_rows, "قدرات ذهنية")


if __name__ == "__main__":
    main()
