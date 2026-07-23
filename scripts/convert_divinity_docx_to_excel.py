# -*- coding: utf-8 -*-
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCX_PATH = os.path.join(BASE, "data", "Divinity_of_Christ_300_Questions.docx")
OUT_PATH = os.path.join(BASE, "output", "Divinity_of_Christ_300_Questions.xlsx")

NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


HEADER_FIRST = "المجموعة الأولى"
HEADER_SECOND = "المجموعة الثانية"
HEADER_THIRD = "المجموعة الثالثة"


def extract_paragraphs(docx_path):
    with zipfile.ZipFile(docx_path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)

    paras = []
    for p in root.findall(".//w:p", NS):
        txt = "".join((n.text or "") for n in p.findall(".//w:t", NS)).strip()
        if txt:
            paras.append(re.sub(r"\s+", " ", txt))
    return paras


def parse_question_group_1_or_3(text):
    m = re.match(r"^س\s*(\d+)\s*:\s*(.*?)\s*الإجابة\s*:\s*(.+)$", text)
    if not m:
        return None
    stem = clean_question_stem(m.group(2).strip())
    ans = m.group(3).strip()
    return stem, ans


def clean_question_stem(stem):
    # Remove trailing source marker like: (سؤال 12)
    stem = re.sub(r"\s*\(\s*سؤال\s*\d+\s*\)\s*", " ", stem)
    return re.sub(r"\s+", " ", stem).strip()


def extract_false_reason(answer_text):
    m = re.match(r"^خطأ\s*[-–]\s*(.+)$", answer_text)
    if m:
        return m.group(1).strip()
    return ""


def parse_question_group_2(text):
    m = re.match(
        r"^س\s*(\d+)\s*:\s*(.*?)\s*أ\)\s*(.*?)\s*ب\)\s*(.*?)\s*ج\)\s*(.*?)\s*الإجابة الصحيحة\s*:\s*([أبج])\)\s*(.+)$",
        text,
    )
    if not m:
        return None
    stem = clean_question_stem(m.group(2).strip())
    opt_a = m.group(3).strip()
    opt_b = m.group(4).strip()
    opt_c = m.group(5).strip()
    ans_letter = m.group(6).strip()
    return stem, opt_a, opt_b, opt_c, ans_letter


def normalize_for_duplicates(stem):
    return re.sub(r"\s+", " ", stem).strip().lower()


def deduplicate_rows(rows):
    seen = set()
    unique_rows = []
    for row in rows:
        key = normalize_for_duplicates(row[1])
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows


def annotate_duplicates(rows):
    by_stem = defaultdict(list)
    for row in rows:
        by_stem[normalize_for_duplicates(row[1])].append(row[0])

    out = []
    for row in rows:
        ids = by_stem[normalize_for_duplicates(row[1])]
        if len(ids) > 1:
            others = [n for n in ids if n != row[0]]
            dup_note = "مكرر مع: " + "، ".join(str(n) for n in others)
        else:
            dup_note = "فريد"
        out.append(row + (dup_note,))
    return out


def build_rows(paragraphs):
    rows = []
    group = None

    for p in paragraphs:
        if HEADER_FIRST in p:
            group = HEADER_FIRST
            continue
        if HEADER_SECOND in p:
            group = HEADER_SECOND
            continue
        if HEADER_THIRD in p:
            group = HEADER_THIRD
            continue

        if not p.startswith("س"):
            continue

        if group in (HEADER_FIRST, HEADER_THIRD):
            parsed = parse_question_group_1_or_3(p)
            if not parsed:
                raise ValueError(f"Could not parse line in {group}: {p}")
            stem, ans_text = parsed
            reason = ""

            if group == HEADER_FIRST:
                if ans_text.startswith("صح"):
                    a, b, c, d = "صح", "خطأ", "", ""
                    answer = "أ"
                elif ans_text.startswith("خطأ"):
                    a, b, c, d = "صح", "خطأ", "", ""
                    answer = "ب"
                    reason = extract_false_reason(ans_text)
                else:
                    a, b, c, d = "", "", "", ""
                    answer = ans_text
            else:
                a, b, c, d = "", "", "", ""
                answer = ans_text

            rows.append((len(rows) + 1, stem, a, b, c, d, answer, group, reason))
            continue

        if group == HEADER_SECOND:
            parsed = parse_question_group_2(p)
            if not parsed:
                raise ValueError(f"Could not parse line in {group}: {p}")
            stem, a, b, c, answer = parsed
            rows.append((len(rows) + 1, stem, a, b, c, "", answer, group, ""))
            continue

        raise ValueError(f"Question found before group header: {p}")

    return rows


def write_excel(rows, out_path):
    wb = openpyxl.Workbook()

    headers = ["رقم السؤال", "السؤال", "أ", "ب", "ج", "د", "الإجابة الصحيحة", "المستوى", "التعليل", "التكرار"]
    hf = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="B0B0B0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center", wrap_text=True)

    sheets = {
        HEADER_FIRST: "صح أم خطأ",
        HEADER_SECOND: "اختر الإجابة الصحيحة",
        HEADER_THIRD: "اكمل",
    }

    grouped_rows = {HEADER_FIRST: [], HEADER_SECOND: [], HEADER_THIRD: []}
    for row in rows:
        grouped_rows[row[7]].append(row)

    # Reuse the default worksheet for the first sheet, then create the others.
    ws = wb.active
    ws.title = sheets[HEADER_FIRST]
    ws.sheet_view.rightToLeft = True

    for group_key in (HEADER_FIRST, HEADER_SECOND, HEADER_THIRD):
        if group_key == HEADER_FIRST:
            current_ws = ws
        else:
            current_ws = wb.create_sheet(title=sheets[group_key])
            current_ws.sheet_view.rightToLeft = True

        current_ws.append(headers)
        for c in current_ws[1]:
            c.fill = hf
            c.font = hfont
            c.alignment = ctr
            c.border = bd

        group_counter = 1
        for r in grouped_rows[group_key]:
            row_values = [group_counter, r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9]]
            current_ws.append(row_values)
            group_counter += 1

        for col, width in {"A": 10, "B": 58, "C": 24, "D": 24, "E": 24, "F": 24, "G": 22, "H": 18, "I": 50, "J": 45}.items():
            current_ws.column_dimensions[col].width = width

        for rr in range(2, current_ws.max_row + 1):
            for col in range(1, 11):
                cell = current_ws.cell(row=rr, column=col)
                cell.border = bd
                cell.alignment = rgt if col in (2, 9, 10) else ctr
            current_ws.row_dimensions[rr].height = 30

        current_ws.freeze_panes = "A2"

    try:
        wb.save(out_path)
    except PermissionError:
        raise SystemExit(f"Cannot write '{out_path}' - please close it in Excel and re-run.")


if __name__ == "__main__":
    paragraphs = extract_paragraphs(DOCX_PATH)
    rows = build_rows(paragraphs)
    if len(rows) != 300:
        raise SystemExit(f"Expected 300 questions, found {len(rows)}")
    rows = deduplicate_rows(rows)
    rows = annotate_duplicates(rows)
    write_excel(rows, OUT_PATH)
    print("Saved", OUT_PATH)
    print("Unique questions:", len(rows))
    print("Total rows (incl. header):", len(rows) + 1)
