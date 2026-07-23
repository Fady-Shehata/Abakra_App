"""Workbook-backed question content loading and normalization.

SQLite never stores question/answer/choice text. This module reads content
from the managed source workbook on demand, with a small in-process cache.
"""
from __future__ import annotations

import hashlib
import threading
from functools import lru_cache
from pathlib import Path
from typing import Optional

import openpyxl

from . import config, models

# Recognized Arabic/English header variations -> canonical field
HEADER_MAP = {
    "question": {"السؤال", "question", "نص السؤال", "q"},
    "a": {"أ", "ا", "choice a", "a", "الاختيار أ", "خيار أ", "option a"},
    "b": {"ب", "choice b", "b", "الاختيار ب", "خيار ب", "option b"},
    "c": {"ج", "choice c", "c", "الاختيار ج", "خيار ج", "option c"},
    "d": {"د", "choice d", "d", "الاختيار د", "خيار د", "option d"},
    "answer": {"الإجابة الصحيحة", "الاجابة الصحيحة", "correct answer", "answer", "الإجابة", "الاجابة"},
    "difficulty": {"المستوى", "difficulty", "level", "الصعوبة"},
    "explanation": {"التعليل", "explanation", "شرح", "note"},
    "number": {"رقم السؤال", "number", "no", "م", "#", "id"},
}

CHOICE_LETTERS = {"أ": 0, "ا": 0, "ب": 1, "ج": 2, "د": 3, "a": 0, "b": 1, "c": 2, "d": 3}

_cache_lock = threading.Lock()


def normalize(text) -> str:
    if text is None:
        return ""
    s = str(text).strip()
    # normalize alef/hamza variants and remove tatweel for hashing/dedup
    for a, b in (("أ", "ا"), ("إ", "ا"), ("آ", "ا"), ("ـ", "")):
        s = s.replace(a, b)
    return " ".join(s.split())


def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def content_hash(question: str, answer: str, choices: list[str]) -> str:
    key = normalize(question) + "|" + normalize(answer) + "|" + "|".join(normalize(c) for c in choices)
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


def map_headers(header_row: list) -> dict:
    """Map column index -> canonical field name."""
    mapping: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        label = normalize(cell).lower()
        for field, variants in HEADER_MAP.items():
            if label in {normalize(v).lower() for v in variants}:
                mapping[idx] = field
                break
    return mapping


def detect_qtype(sheet_name: str, choices: list[str], answer: str) -> str:
    name = normalize(sheet_name).lower()
    non_empty = [c for c in choices if c]
    if "صح" in name and "خطا" in name:
        return "tf"
    if len(non_empty) == 2 and {normalize(c) for c in non_empty} <= {"صح", "خطا", "true", "false"}:
        return "tf"
    if len(non_empty) >= 2:
        return "mc"
    return "open"


@lru_cache(maxsize=64)
def _load_workbook_rows(stored_path: str, mtime: float) -> dict:
    """Return {sheet_name: {row_number: dict}} for the whole workbook.

    mtime participates in the cache key so edits invalidate automatically.
    """
    wb = openpyxl.load_workbook(stored_path, read_only=True, data_only=True)
    result: dict[str, dict[int, dict]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[ws.title] = {}
            continue
        header = list(rows[0])
        mapping = map_headers(header)
        sheet_rows: dict[int, dict] = {}
        for r_i, raw in enumerate(rows[1:], start=2):
            fields = {"a": "", "b": "", "c": "", "d": "", "question": "",
                      "answer": "", "difficulty": "", "explanation": "", "number": ""}
            for idx, field in mapping.items():
                if idx < len(raw):
                    fields[field] = "" if raw[idx] is None else str(raw[idx]).strip()
            sheet_rows[r_i] = fields
        result[ws.title] = sheet_rows
    wb.close()
    return result


def _stored_path(source: models.QuestionSource) -> Path:
    return config.WORKBOOK_STORE / source.stored_filename


def invalidate_cache() -> None:
    with _cache_lock:
        _load_workbook_rows.cache_clear()


def load_row(source: models.QuestionSource, sheet: str, row: int) -> Optional[dict]:
    path = _stored_path(source)
    if not path.exists():
        return None
    mtime = path.stat().st_mtime
    with _cache_lock:
        data = _load_workbook_rows(str(path), mtime)
    return data.get(sheet, {}).get(row)


def resolve_choices(fields: dict) -> list[str]:
    return [fields.get("a", ""), fields.get("b", ""), fields.get("c", ""), fields.get("d", "")]


def display_answer(fields: dict, qtype: str) -> str:
    """Resolve the human-readable correct answer text."""
    ans = (fields.get("answer") or "").strip()
    if qtype in ("mc", "tf"):
        letter = normalize(ans).lower()
        if letter in CHOICE_LETTERS:
            choices = resolve_choices(fields)
            idx = CHOICE_LETTERS[letter]
            if idx < len(choices) and choices[idx]:
                return choices[idx]
    return ans


def render_question(db, question: models.Question, include_answer: bool = False) -> dict:
    """Return safe, escaped-ready content for a question. Missing workbook is
    handled gracefully."""
    source = db.get(models.QuestionSource, question.source_id)
    if source is None:
        return {"error": "missing_source", "code": question.question_code}
    fields = load_row(source, question.worksheet, question.row_number)
    if fields is None:
        return {"error": "missing_workbook", "code": question.question_code}
    choices = [c for c in resolve_choices(fields) if c]
    out = {
        "id": question.id,
        "code": question.question_code,
        "qtype": question.qtype,
        "text": fields.get("question", ""),
        "choices": choices,
        "difficulty": question.difficulty,
    }
    if include_answer:
        out["answer"] = display_answer(fields, question.qtype)
        out["explanation"] = fields.get("explanation", "")
    return out
