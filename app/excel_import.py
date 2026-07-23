"""Multi-sheet Excel import for questions and teams.

Resilient: an error in one worksheet/row never aborts the whole import.
Copies uploaded workbooks into managed storage (immutable source).
"""
from __future__ import annotations

import datetime as dt
import json
import re
import shutil
import time
import uuid
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from . import config, models, question_store as qs


# --------------------------------------------------------------------------- #
# Safe storage
# --------------------------------------------------------------------------- #
def safe_stored_name(original: str, suffix: str = ".xlsx") -> str:
    stem = re.sub(r"[^A-Za-z0-9_\-]", "_", Path(original).stem)[:60]
    return f"{stem}_{uuid.uuid4().hex[:12]}{suffix}"


def store_workbook(temp_path: Path, original_name: str) -> tuple[Path, str, str]:
    stored = safe_stored_name(original_name)
    dest = config.WORKBOOK_STORE / stored
    shutil.copyfile(temp_path, dest)
    return dest, stored, qs.file_hash(dest)


# --------------------------------------------------------------------------- #
# Question import
# --------------------------------------------------------------------------- #
def import_questions_workbook(
    db: Session,
    temp_path: Path,
    original_name: str,
    category: models.Category,
    user_id: Optional[int] = None,
) -> dict:
    started = time.time()
    summary = {
        "workbook_name": original_name,
        "worksheets_detected": 0,
        "worksheets_processed": 0,
        "worksheets_with_warnings": 0,
        "worksheets_with_errors": 0,
        "rows_processed": 0,
        "questions_imported": 0,
        "duplicates_skipped": 0,
        "invalid_rows": 0,
        "warnings": [],
        "errors": [],
        "duration_seconds": 0.0,
    }

    ext = Path(original_name).suffix.lower()
    if ext not in config.ALLOWED_EXCEL_EXT:
        summary["errors"].append("unsupported_file")
        summary["duration_seconds"] = round(time.time() - started, 3)
        return summary

    try:
        # keep_vba disabled -> macros never executed
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
    except Exception:
        summary["errors"].append("corrupted_file")
        summary["duration_seconds"] = round(time.time() - started, 3)
        return summary

    # Copy to managed storage & register source
    dest, stored, fhash = store_workbook(temp_path, original_name)
    source = models.QuestionSource(
        original_name=original_name, stored_filename=stored, file_hash=fhash, kind="import"
    )
    db.add(source)
    db.flush()

    # existing hashes for dedup across previous imports
    existing_hashes = {h for (h,) in db.query(models.Question.content_hash).all()}
    imported_now: set[str] = set()

    summary["worksheets_detected"] = len(wb.worksheets)
    seq = db.query(models.Question).count()

    for ws in wb.worksheets:
        sheet_warn = False
        sheet_err = False
        try:
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            summary["worksheets_with_errors"] += 1
            summary["errors"].append(f"{ws.title}: read_error")
            continue
        if not rows or len(rows) < 2:
            summary["warnings"].append(f"{ws.title}: empty")
            sheet_warn = True
            summary["worksheets_with_warnings"] += 1
            continue

        header = list(rows[0])
        mapping = qs.map_headers(header)
        if "question" not in mapping.values():
            summary["errors"].append(f"{ws.title}: no_question_column")
            summary["worksheets_with_errors"] += 1
            continue

        for r_i, raw in enumerate(rows[1:], start=2):
            summary["rows_processed"] += 1
            fields = {"a": "", "b": "", "c": "", "d": "", "question": "",
                      "answer": "", "difficulty": "", "explanation": ""}
            for idx, field in mapping.items():
                if field in fields and idx < len(raw) and raw[idx] is not None:
                    val = raw[idx]
                    # Ignore formula strings safely (never evaluate)
                    fields[field] = str(val).strip()

            question_text = fields["question"]
            if not question_text:
                summary["invalid_rows"] += 1
                continue

            choices = qs.resolve_choices(fields)
            qtype = qs.detect_qtype(ws.title, choices, fields["answer"])

            # validation by type
            answer = fields["answer"].strip()
            non_empty_choices = [c for c in choices if c]
            if qtype == "mc" and (len(non_empty_choices) < 2 or not answer):
                summary["invalid_rows"] += 1
                sheet_warn = True
                continue
            if qtype in ("tf", "open") and not answer:
                summary["invalid_rows"] += 1
                sheet_warn = True
                continue

            chash = qs.content_hash(question_text, answer, choices)
            if chash in existing_hashes or chash in imported_now:
                summary["duplicates_skipped"] += 1
                continue

            seq += 1
            code = f"{category.id:02d}-{source.id:04d}-{seq:05d}"
            q = models.Question(
                category_id=category.id,
                source_id=source.id,
                worksheet=ws.title,
                row_number=r_i,
                question_code=code,
                content_hash=chash,
                qtype=qtype,
                difficulty=fields["difficulty"] or None,
                is_active=True,
            )
            db.add(q)
            imported_now.add(chash)
            summary["questions_imported"] += 1

        if sheet_err:
            summary["worksheets_with_errors"] += 1
        elif sheet_warn:
            summary["worksheets_with_warnings"] += 1
        summary["worksheets_processed"] += 1

    wb.close()
    summary["duration_seconds"] = round(time.time() - started, 3)

    db.add(models.QuestionImport(
        source_id=source.id, category_id=category.id, user_id=user_id,
        summary_json=json.dumps(summary, ensure_ascii=False),
    ))
    db.commit()
    qs.invalidate_cache()
    return summary


# --------------------------------------------------------------------------- #
# Team import
# --------------------------------------------------------------------------- #
TEAM_NAME_HEADERS = {"اسم الفريق", "الفريق", "team", "team name", "name", "الاسم"}
MEMBER_HEADERS = {"الأعضاء", "الاعضاء", "members", "member", "عضو", "أعضاء الفريق"}
COUNT_HEADERS = {"عدد الأعضاء", "عدد الاعضاء", "member count", "count", "العدد"}
LEVEL_HEADERS = {"المرحلة", "المستوى", "level", "stage", "grade", "المرحله"}


def _match_header(label: str, options: set[str]) -> bool:
    return qs.normalize(label).lower() in {qs.normalize(o).lower() for o in options}


def import_teams_workbook(db: Session, temp_path: Path, original_name: str) -> dict:
    started = time.time()
    summary = {
        "workbook_name": original_name,
        "rows_processed": 0,
        "teams_imported": 0,
        "duplicate_teams_skipped": 0,
        "invalid_rows": 0,
        "warnings": [],
        "errors": [],
        "duration_seconds": 0.0,
    }
    ext = Path(original_name).suffix.lower()
    if ext not in config.ALLOWED_EXCEL_EXT:
        summary["errors"].append("unsupported_file")
        return summary
    try:
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
    except Exception:
        summary["errors"].append("corrupted_file")
        return summary

    existing = {qs.normalize(n).lower() for (n,) in db.query(models.Team.name).all()}
    seen: set[str] = set()

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue
        header = list(rows[0])
        name_idx = member_idx = count_idx = None
        level_idx: int | None = None
        member_cols: list[int] = []
        for idx, cell in enumerate(header):
            if cell is None:
                continue
            label = str(cell)
            if name_idx is None and _match_header(label, TEAM_NAME_HEADERS):
                name_idx = idx
            elif _match_header(label, MEMBER_HEADERS):
                member_idx = idx
                member_cols.append(idx)
            elif count_idx is None and _match_header(label, COUNT_HEADERS):
                count_idx = idx
            elif level_idx is None and _match_header(label, LEVEL_HEADERS):
                level_idx = idx
        # also treat columns like "عضو 1", "member 1" as member columns
        for idx, cell in enumerate(header):
            if cell is None or idx in member_cols:
                continue
            label = qs.normalize(cell).lower()
            if re.match(r"^(عضو|member|player)\s*\d+$", label):
                member_cols.append(idx)

        if name_idx is None:
            summary["errors"].append(f"{ws.title}: no_team_name_column")
            continue

        for raw in rows[1:]:
            summary["rows_processed"] += 1
            name = "" if name_idx >= len(raw) or raw[name_idx] is None else str(raw[name_idx]).strip()
            if not name:
                summary["invalid_rows"] += 1
                continue
            key = qs.normalize(name).lower()
            if key in existing or key in seen:
                summary["duplicate_teams_skipped"] += 1
                continue

            members: list[str] = []
            for mc in member_cols:
                if mc < len(raw) and raw[mc]:
                    # a single cell may contain comma/newline separated names
                    for part in re.split(r"[،,\n;]+", str(raw[mc])):
                        p = part.strip()
                        if p and p not in members:
                            members.append(p)

            level_val: str | None = None
            if level_idx is not None and level_idx < len(raw) and raw[level_idx] is not None:
                lv = str(raw[level_idx]).strip()
                if lv:
                    level_val = lv

            team = models.Team(name=name, level=level_val,
                               member_count=len(members), is_active=True)
            for m in members:
                team.members.append(models.TeamMember(name=m))
            db.add(team)
            seen.add(key)
            summary["teams_imported"] += 1

    wb.close()
    db.commit()
    summary["duration_seconds"] = round(time.time() - started, 3)
    return summary
