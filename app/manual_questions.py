"""Manual question entry stored in application-managed workbooks.

Question text is never stored in SQLite. Manual questions are appended to a
per-category manual workbook; SQLite keeps only references and metadata.
"""
from __future__ import annotations

import openpyxl
from pathlib import Path
from sqlalchemy.orm import Session

from . import config, models, question_store as qs

HEADER = ["رقم السؤال", "السؤال", "أ", "ب", "ج", "د", "الإجابة الصحيحة", "المستوى", "التعليل"]


def _manual_source(db: Session, category: models.Category) -> models.QuestionSource:
    stored = f"manual_cat_{category.id}.xlsx"
    source = (
        db.query(models.QuestionSource)
        .filter(models.QuestionSource.stored_filename == stored)
        .first()
    )
    path = config.WORKBOOK_STORE / stored
    if source is None or not path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "يدوي"
        ws.append(HEADER)
        wb.save(path)
        wb.close()
        if source is None:
            source = models.QuestionSource(
                original_name=f"manual_{category.name}.xlsx",
                stored_filename=stored,
                file_hash=qs.file_hash(path),
                kind="manual",
            )
            db.add(source)
            db.flush()
    return source


def add_manual_question(
    db: Session,
    category: models.Category,
    text: str,
    qtype: str,
    choices: list[str],
    answer: str,
    explanation: str = "",
    difficulty: str = "",
) -> models.Question:
    source = _manual_source(db, category)
    path = config.WORKBOOK_STORE / source.stored_filename
    wb = openpyxl.load_workbook(path)
    ws = wb["يدوي"]
    row_number = ws.max_row + 1
    a, b, c, d = (choices + ["", "", "", ""])[:4]
    ws.append([row_number - 1, text, a, b, c, d, answer, difficulty, explanation])
    wb.save(path)
    wb.close()

    source.file_hash = qs.file_hash(path)
    chash = qs.content_hash(text, answer, choices)
    seq = db.query(models.Question).count() + 1
    code = f"{category.id:02d}-{source.id:04d}-M{seq:05d}"
    q = models.Question(
        category_id=category.id,
        source_id=source.id,
        worksheet="يدوي",
        row_number=row_number,
        question_code=code,
        content_hash=chash,
        qtype=qtype,
        difficulty=difficulty or None,
        is_active=True,
    )
    db.add(q)
    db.commit()
    qs.invalidate_cache()
    return q


def edit_manual_question(
    db: Session,
    question: models.Question,
    text: str,
    qtype: str,
    choices: list[str],
    answer: str,
    explanation: str = "",
    difficulty: str = "",
) -> models.Question:
    """Create a new version row and deactivate the old reference."""
    category = db.get(models.Category, question.category_id)
    question.is_active = False
    db.flush()
    return add_manual_question(db, category, text, qtype, choices, answer, explanation, difficulty)
