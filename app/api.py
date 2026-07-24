"""JSON API for the live game screen, reports export and usage reset."""
from __future__ import annotations

import io
import json
import datetime as dt
import openpyxl

from fastapi import APIRouter, Depends, Request, Body, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

from . import config, models, security, game as ge, question_store as qs, scoring, match_service
from .database import get_db
from .security import require_login, require_admin

router = APIRouter()


def _get_match_for_user(mid: int, request: Request, db: Session) -> models.Match:
    user = security.get_current_user(request, db)
    if not user:
        raise HTTPException(401, "login_required")
    m = db.get(models.Match, mid)
    if not m:
        raise HTTPException(404, "not_found")
    if not security.can_access_match(user, m):
        raise HTTPException(403, "forbidden")
    return m


def _serialize_state(db: Session, match: models.Match) -> dict:
    session = match.session
    state = ge.load_state(session) if session else ge._default_state()
    remaining = ge.remaining_by_category(db, session.id) if session else []
    section_names = scoring.section_names(db)
    section_types = scoring.section_types(db)
    current = state.get("current")
    current_view = None
    if current:
        q = db.get(models.Question, current["question_id"])
        # Only expose text once revealed
        include = current["phase"] in ("revealed", "rebound_open", "done")
        content = qs.render_question(db, q, include_answer=include) if q else {}
        if not include:
            content.pop("text", None)
            content.pop("choices", None)
        current_view = {
            **current,
            "section_type": current.get("section_type") or section_types.get(current.get("section")),
            "content": content,
        }
    # history
    events = (
        db.query(models.ScoreEvent)
        .filter(models.ScoreEvent.match_id == match.id)
        .order_by(models.ScoreEvent.created_at.desc())
        .limit(30)
        .all()
    )
    history = [{
        "team": ("a" if e.team_id == match.team_a_id else "b" if e.team_id == match.team_b_id else None),
        "delta": e.delta, "reason": e.reason, "section": e.section,
        "at": e.created_at.isoformat(),
    } for e in events]
    special = db.query(models.Category).filter_by(is_regular=False).first()
    return {
        "match_id": match.id,
        "status": match.status,
        "stage": match.stage,
        "team_a": {"id": match.team_a_id, "name": match.team_a.name if match.team_a else "?",
                   "logo": match.team_a.logo_path if match.team_a else None},
        "team_b": {"id": match.team_b_id, "name": match.team_b.name if match.team_b else "?",
                   "logo": match.team_b.logo_path if match.team_b else None},
        "score_a": match.score_a,
        "score_b": match.score_b,
        "current_section": state.get("current_section", 0),
        "sections": state.get("sections", {}),
        "current": current_view,
        "wheel": state.get("wheel"),
        "buzzer": state.get("buzzer"),
        "last_spin": state.get("last_spin"),
        "remaining": remaining,
        "section_names": section_names,
        "section_order": scoring.section_order(db),
        "section_types": section_types,
        "history": history,
        "winner_team_id": match.winner_team_id,
        "is_draw": match.is_draw,
        "special_category_id": special.id if special else None,
    }


@router.get("/game/{mid}/state")
def game_state(mid: int, request: Request, db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    return _serialize_state(db, m)


def _require_session(match: models.Match) -> models.GameSession:
    if not match.session:
        raise HTTPException(400, "no_session")
    return match.session


@router.post("/game/{mid}/start-section")
def start_section(mid: int, request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    s = _require_session(m)
    try:
        ge.start_section(db, s, int(payload.get("section")))
    except ge.GameError as e:
        return JSONResponse({"error": e.key, "ctx": e.ctx}, status_code=400)
    return _serialize_state(db, m)


@router.post("/game/{mid}/select")
def select(mid: int, request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    s = _require_session(m)
    user = security.get_current_user(request, db)
    try:
        ge.select_question(db, s, int(payload["section"]), int(payload["category_id"]),
                           payload.get("team"), user.id if user else None,
                           via_joker=bool(payload.get("via_joker")))
    except ge.GameError as e:
        return JSONResponse({"error": e.key, "ctx": e.ctx}, status_code=400)
    return _serialize_state(db, m)


@router.post("/game/{mid}/reveal")
def reveal(mid: int, request: Request, db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    s = _require_session(m)
    try:
        ge.reveal(db, s)
    except ge.GameError as e:
        return JSONResponse({"error": e.key}, status_code=400)
    return _serialize_state(db, m)


@router.post("/game/{mid}/buzz")
def buzz(mid: int, request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    ge.set_buzz(db, _require_session(m), payload.get("team"))
    return _serialize_state(db, m)


@router.post("/game/{mid}/reset-buzzer")
def reset_buzzer(mid: int, request: Request, db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    ge.reset_buzzer(db, _require_session(m))
    return _serialize_state(db, m)


@router.post("/game/{mid}/mark")
def mark(mid: int, request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    s = _require_session(m)
    user = security.get_current_user(request, db)
    hid = user.id if user else None
    action = payload.get("action")
    try:
        if action == "a_correct":
            ge.mark_correct(db, s, "a", hid)
        elif action == "b_correct":
            ge.mark_correct(db, s, "b", hid)
        elif action == "wrong":
            ge.mark_wrong(db, s, hid)
        elif action == "open_rebound":
            ge.open_rebound(db, s, hid)
        elif action == "rebound_correct":
            ge.rebound_correct(db, s, hid)
        elif action == "rebound_wrong":
            ge.rebound_wrong(db, s, hid)
        elif action == "father_a":
            ge.father_award(db, s, "a", hid)
        elif action == "father_b":
            ge.father_award(db, s, "b", hid)
        elif action == "father_none":
            ge.father_award(db, s, None, hid)
        elif action == "skip":
            ge.skip_question(db, s, hid)
        elif action == "invalidate":
            ge.invalidate_question(db, s, payload.get("reason", ""), hid)
        else:
            return JSONResponse({"error": "invalid_transition"}, status_code=400)
    except ge.GameError as e:
        return JSONResponse({"error": e.key, "ctx": e.ctx}, status_code=400)
    return _serialize_state(db, m)


@router.post("/game/{mid}/spin")
def spin(mid: int, request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    s = _require_session(m)
    try:
        ge.spin_wheel(db, s, payload.get("team"))
    except ge.GameError as e:
        return JSONResponse({"error": e.key}, status_code=400)
    return _serialize_state(db, m)


@router.post("/game/{mid}/finish-section")
def finish_section(mid: int, request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    ge.finish_section(db, _require_session(m), int(payload.get("section")))
    return _serialize_state(db, m)


@router.post("/game/{mid}/pause")
def pause(mid: int, request: Request, db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    m.status = "paused"
    if m.session:
        m.session.status = "paused"
    db.commit()
    return _serialize_state(db, m)


@router.post("/game/{mid}/start")
def start_match(mid: int, request: Request, db: Session = Depends(get_db)):
    """Flip a scheduled/ready match to in_progress. Explicit host action so
    opening the page never auto-starts the game."""
    m = _get_match_for_user(mid, request, db)
    if m.status in ("scheduled", "ready"):
        m.status = "in_progress"
        if not m.started_at:
            m.started_at = dt.datetime.utcnow()
    if m.session and m.session.status != "in_progress":
        m.session.status = "in_progress"
    db.commit()
    return _serialize_state(db, m)


@router.post("/game/{mid}/reset-to-ready")
def reset_to_ready(mid: int, request: Request, db: Session = Depends(get_db),
                   user: models.User = Depends(require_login)):
    """Return an in-progress or paused match to the ready state so it can be
    started fresh. Only permitted while the match is not yet completed."""
    m = db.get(models.Match, mid)
    if not m:
        raise HTTPException(404, "not_found")
    if not security.can_access_match(user, m):
        raise HTTPException(403, "forbidden")
    if m.status == "completed":
        return JSONResponse({"error": "match_completed"}, status_code=400)
    m.status = "ready"
    m.started_at = None
    if m.session:
        m.session.status = "ready"
    db.commit()
    security.audit(db, user.id, "reset_to_ready", f"match={mid}")
    return _serialize_state(db, m)


@router.post("/game/{mid}/resume")
def resume(mid: int, request: Request, db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    m.status = "in_progress"
    if m.session:
        m.session.status = "in_progress"
    db.commit()
    return _serialize_state(db, m)


@router.post("/game/{mid}/complete")
def complete(mid: int, request: Request, payload: dict = Body(default={}), db: Session = Depends(get_db)):
    m = _get_match_for_user(mid, request, db)
    user = security.get_current_user(request, db)
    try:
        match_service.complete_match(db, m, user.id if user else None,
                                     forced_winner_side=payload.get("winner_side"))
    except ge.GameError as e:
        return JSONResponse({"error": e.key}, status_code=400)
    return _serialize_state(db, m)


@router.post("/game/{mid}/reset-usage")
def reset_usage(mid: int, request: Request, db: Session = Depends(get_db),
                user: models.User = Depends(require_admin)):
    m = db.get(models.Match, mid)
    if not m or not m.session:
        raise HTTPException(404, "not_found")
    db.query(models.QuestionUsage).filter_by(session_id=m.session.id).delete()
    m.session.state_json = None
    m.session.current_section = 0
    db.commit()
    security.audit(db, user.id, "reset_usage", f"match={mid}")
    return {"ok": True}


# --------------------------------------------------------------------------- #
# Reports export
# --------------------------------------------------------------------------- #
def _xlsx_response(rows: list[list], filename: str) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "report"
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reports/teams.xlsx")
def report_teams_xlsx(request: Request, db: Session = Depends(get_db),
                      user: models.User = Depends(require_admin)):
    rows = [["id", "name", "level", "members", "active"]]
    for t in db.query(models.Team).order_by(models.Team.name).all():
        rows.append([t.id, t.name, t.level or "", "؛ ".join(m.name for m in t.members), "1" if t.is_active else "0"])
    return _xlsx_response(rows, "teams.xlsx")


@router.get("/reports/standings/{tid}.xlsx")
def report_standings_xlsx(tid: int, db: Session = Depends(get_db),
                          user: models.User = Depends(require_admin)):
    from . import standings as st
    t = db.get(models.Tournament, tid)
    wb = openpyxl.Workbook()
    header = ["rank", "team", "P", "W", "D", "L", "Pts", "GF", "GA", "GD"]

    def safe_sheet_name(name: str, fallback: str) -> str:
        cleaned = "".join("_" if ch in r'[]:*?/\\' else ch for ch in (name or fallback)).strip()
        return (cleaned or fallback)[:31]

    if t and t.groups:
        wb.remove(wb.active)
        used_names: set[str] = set()
        for idx, g in enumerate(t.groups, start=1):
            base_name = safe_sheet_name(g.name, f"group_{idx}")
            sheet_name = base_name
            suffix = 2
            while sheet_name in used_names:
                tail = f"_{suffix}"
                sheet_name = f"{base_name[:31 - len(tail)]}{tail}"
                suffix += 1
            used_names.add(sheet_name)
            ws = wb.create_sheet(sheet_name)
            ws.append(header)
            for r in st.compute_group_standings(db, t, g):
                ws.append([r["rank"], r["team_name"], r["played"], r["won"],
                           r["drawn"], r["lost"], r["points"], r["gf"], r["ga"], r["gd"]])
    else:
        ws = wb.active
        ws.title = "standings"
        ws.append(header)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=standings_{tid}.xlsx"},
    )


@router.get("/reports/results/{tid}.xlsx")
def report_results_xlsx(tid: int, db: Session = Depends(get_db),
                        user: models.User = Depends(require_admin)):
    rows = [["stage", "round", "team_a", "team_b", "score_a", "score_b", "winner", "host"]]
    matches = db.query(models.Match).filter_by(tournament_id=tid).all()
    for m in matches:
        winner = m.winner_team_id and (m.team_a.name if m.winner_team_id == m.team_a_id else m.team_b.name)
        rows.append([m.stage, m.round_name or "",
                     m.team_a.name if m.team_a else "", m.team_b.name if m.team_b else "",
                     m.score_a, m.score_b, winner or ("تعادل" if m.is_draw else ""),
                     m.host.display_name if m.host else ""])
    return _xlsx_response(rows, f"results_{tid}.xlsx")


def _group_matches_rows(db: Session, tid: int) -> list[list]:
    """Detailed export of every match played in the group stage of a
    tournament. Ordered by group name then scheduled time so the file is
    readable as a tournament schedule + result sheet."""
    header = [
        "group", "round", "scheduled_at",
        "team_a", "team_b", "host", "status",
        "score_a", "score_b", "winner", "started_at", "completed_at",
    ]
    rows: list[list] = [header]
    q = (
        db.query(models.Match)
        .filter(models.Match.tournament_id == tid, models.Match.stage == "group")
        .order_by(
            models.Match.group_id.is_(None),
            models.Match.group_id.asc(),
            models.Match.scheduled_at.is_(None),
            models.Match.scheduled_at.asc(),
            models.Match.created_at.asc(),
        )
    )
    groups_by_id = {g.id: g.name for g in db.query(models.Group).filter_by(tournament_id=tid).all()}
    for m in q.all():
        gname = groups_by_id.get(m.group_id or 0, "")
        winner = ""
        if m.is_draw:
            winner = "تعادل"
        elif m.winner_team_id:
            winner = (m.team_a.name if m.winner_team_id == m.team_a_id
                      else m.team_b.name if m.winner_team_id == m.team_b_id else "")
        rows.append([
            gname,
            m.round_name or "",
            m.scheduled_at.strftime("%Y-%m-%d %H:%M") if m.scheduled_at else "",
            m.team_a.name if m.team_a else "",
            m.team_b.name if m.team_b else "",
            m.host.display_name if m.host else "",
            m.status,
            m.score_a,
            m.score_b,
            winner,
            m.started_at.strftime("%Y-%m-%d %H:%M") if m.started_at else "",
            m.completed_at.strftime("%Y-%m-%d %H:%M") if m.completed_at else "",
        ])
    return rows


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names are max 31 chars and disallow \\/?*[]: — sanitize and
    de-duplicate against sheet names already used in the same workbook."""
    if not name:
        name = "Sheet"
    for ch in "\\/?*[]:":
        name = name.replace(ch, " ")
    name = name.strip()[:31] or "Sheet"
    base = name
    i = 2
    while name in used:
        suffix = f" ({i})"
        name = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(name)
    return name


@router.get("/reports/group-matches/{tid}.xlsx")
def report_group_matches_xlsx(tid: int, db: Session = Depends(get_db),
                              user: models.User = Depends(require_admin)):
    """Build a workbook with one sheet per group. Each sheet is titled with
    the group's name and contains only that group's matches. Matches that
    have no group (rare) are collected on a fallback sheet."""
    header = [
        "round", "scheduled_at", "team_a", "team_b", "host", "status",
        "score_a", "score_b", "winner", "started_at", "completed_at",
    ]
    wb = openpyxl.Workbook()
    # Remove the auto-created default sheet; we'll add named ones per group.
    default_ws = wb.active
    wb.remove(default_ws)

    used_names: set[str] = set()
    groups = (
        db.query(models.Group)
        .filter_by(tournament_id=tid)
        .order_by(models.Group.name.asc())
        .all()
    )

    def _append_matches(ws, matches):
        ws.append(header)
        for m in matches:
            winner = ""
            if m.is_draw:
                winner = "تعادل"
            elif m.winner_team_id:
                winner = (m.team_a.name if m.winner_team_id == m.team_a_id
                          else m.team_b.name if m.winner_team_id == m.team_b_id else "")
            ws.append([
                m.round_name or "",
                m.scheduled_at.strftime("%Y-%m-%d %H:%M") if m.scheduled_at else "",
                m.team_a.name if m.team_a else "",
                m.team_b.name if m.team_b else "",
                m.host.display_name if m.host else "",
                m.status,
                m.score_a,
                m.score_b,
                winner,
                m.started_at.strftime("%Y-%m-%d %H:%M") if m.started_at else "",
                m.completed_at.strftime("%Y-%m-%d %H:%M") if m.completed_at else "",
            ])

    for g in groups:
        ws = wb.create_sheet(title=_sanitize_sheet_name(g.name, used_names))
        matches = (
            db.query(models.Match)
            .filter(
                models.Match.tournament_id == tid,
                models.Match.stage == "group",
                models.Match.group_id == g.id,
            )
            .order_by(
                models.Match.scheduled_at.is_(None),
                models.Match.scheduled_at.asc(),
                models.Match.created_at.asc(),
            )
            .all()
        )
        _append_matches(ws, matches)

    # Group-stage matches with no group assignment go on a fallback sheet.
    unassigned = (
        db.query(models.Match)
        .filter(
            models.Match.tournament_id == tid,
            models.Match.stage == "group",
            models.Match.group_id.is_(None),
        )
        .order_by(
            models.Match.scheduled_at.is_(None),
            models.Match.scheduled_at.asc(),
            models.Match.created_at.asc(),
        )
        .all()
    )
    if unassigned:
        ws = wb.create_sheet(title=_sanitize_sheet_name("بدون مجموعة", used_names))
        _append_matches(ws, unassigned)

    # Workbook must have at least one sheet even if the tournament has none.
    if not wb.sheetnames:
        ws = wb.create_sheet(title="report")
        ws.append(header)

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=group_matches_{tid}.xlsx"},
    )


@router.get("/reports/usage/{tid}.xlsx")
def report_usage_xlsx(tid: int, db: Session = Depends(get_db),
                      user: models.User = Depends(require_admin)):
    rows = [["match", "section", "category", "question_code", "state", "result", "rebound", "points"]]
    matches = db.query(models.Match).filter_by(tournament_id=tid).all()
    for m in matches:
        if not m.session:
            continue
        usages = db.query(models.QuestionUsage).filter_by(session_id=m.session.id).all()
        for u in usages:
            cat = db.get(models.Category, u.category_id)
            q = db.get(models.Question, u.question_id)
            rows.append([m.id, u.section, cat.name if cat else "", q.question_code if q else "",
                         u.state, u.answer_result or "", u.rebound_result or "", u.points_awarded])
    return _xlsx_response(rows, f"usage_{tid}.xlsx")
