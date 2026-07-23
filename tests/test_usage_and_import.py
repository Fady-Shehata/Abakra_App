from __future__ import annotations

from pathlib import Path

import openpyxl

from app import excel_import, game as ge, models
from app.database import SessionLocal
from tests.conftest import make_category, make_match_with_session, make_question, make_source


def _clear_categories(db):
    db.query(models.Category).delete()
    db.commit()


def test_question_not_reused_same_session(db_session):
    _clear_categories(db_session)
    cat = make_category(db_session, "كتاب لاهوت", True, True)
    src = make_source(db_session, "u1.xlsx")
    make_question(db_session, cat.id, src.id, "Q1", "h_u1")
    m = make_match_with_session(db_session)

    ge.start_section(db_session, m.session, 2)
    ge.select_question(db_session, m.session, 2, cat.id, None, None)
    ge.reveal(db_session, m.session)
    ge.mark_correct(db_session, m.session, "a", None)

    try:
        ge.select_question(db_session, m.session, 2, cat.id, None, None)
        assert False, "expected no available questions"
    except ge.GameError as e:
        assert e.key in ("not_enough_questions", "question_already_used")


def test_sessions_independent_usage(db_session):
    _clear_categories(db_session)
    cat = make_category(db_session, "كتاب لاهوت", True, True)
    src = make_source(db_session, "u2.xlsx")
    make_question(db_session, cat.id, src.id, "Q1", "h_u2")

    m1 = make_match_with_session(db_session)
    m2 = make_match_with_session(db_session)

    ge.start_section(db_session, m1.session, 2)
    ge.select_question(db_session, m1.session, 2, cat.id, None, None)
    ge.reveal(db_session, m1.session)
    ge.mark_correct(db_session, m1.session, "a", None)

    ge.start_section(db_session, m2.session, 2)
    ge.select_question(db_session, m2.session, 2, cat.id, None, None)
    ge.reveal(db_session, m2.session)
    ge.mark_correct(db_session, m2.session, "b", None)

    assert m1.score_a == 5
    assert m2.score_b == 5


def test_concurrent_select_cannot_double_reserve(db_session):
    _clear_categories(db_session)
    cat = make_category(db_session, "كتاب لاهوت", True, True)
    src = make_source(db_session, "u3.xlsx")
    make_question(db_session, cat.id, src.id, "Q1", "h_u3")
    m = make_match_with_session(db_session)

    ge.start_section(db_session, m.session, 2)

    db2 = SessionLocal()
    try:
        ge.select_question(db_session, m.session, 2, cat.id, None, None)
        m2 = db2.get(models.Match, m.id)
        try:
            ge.select_question(db2, m2.session, 2, cat.id, None, None)
            assert False, "expected reservation conflict"
        except ge.GameError as e:
            assert e.key in ("not_enough_questions", "question_already_used", "invalid_transition")
    finally:
        db2.close()


def test_multisheet_import_with_hidden_and_invalid_sheet(db_session, tmp_path: Path):
    wb_path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Main"
    ws1.append(["رقم السؤال", "السؤال", "أ", "ب", "الإجابة الصحيحة", "المستوى"])
    ws1.append([1, "سؤال 1", "صح", "خطأ", "أ", "سهل"])
    ws2 = wb.create_sheet("HiddenGood")
    ws2.sheet_state = "hidden"
    ws2.append(["Question", "Choice A", "Choice B", "Correct Answer"])
    ws2.append(["Q2", "x", "y", "A"])
    ws3 = wb.create_sheet("Bad")
    ws3.append(["X", "Y", "Z"])
    ws3.append([1, 2, 3])
    wb.save(wb_path)
    wb.close()

    cat = make_category(db_session, "كتاب لاهوت", True, True)
    summary = excel_import.import_questions_workbook(db_session, wb_path, wb_path.name, cat, None)

    assert summary["worksheets_detected"] == 3
    assert summary["questions_imported"] >= 2
    assert summary["worksheets_with_errors"] >= 1


def test_import_corrupted_workbook(db_session, tmp_path: Path):
    bad = tmp_path / "bad.xlsx"
    bad.write_text("not an excel", encoding="utf-8")
    cat = make_category(db_session, "كتاب لاهوت", True, True)
    summary = excel_import.import_questions_workbook(db_session, bad, bad.name, cat, None)
    assert "corrupted_file" in summary["errors"]
