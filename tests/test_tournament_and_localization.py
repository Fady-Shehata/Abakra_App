from __future__ import annotations

import io

import openpyxl

from app import brackets, i18n, match_service, models, services, standings
from tests.conftest import make_team


def test_group_standings_points_and_order(db_session):
    t = models.Tournament(name="T1", status="active")
    db_session.add(t)
    db_session.commit()
    db_session.refresh(t)

    ta = make_team(db_session, "A")
    tb = make_team(db_session, "B")
    tc = make_team(db_session, "C")
    for team in (ta, tb, tc):
        db_session.add(models.TournamentTeam(tournament_id=t.id, team_id=team.id))
    g = models.Group(tournament_id=t.id, name="G1")
    db_session.add(g)
    db_session.commit()
    db_session.refresh(g)
    for team in (ta, tb, tc):
        db_session.add(models.GroupTeam(group_id=g.id, team_id=team.id))
    db_session.commit()

    # A beat B, B drew C, A beat C => A top with 6
    db_session.add(models.Match(tournament_id=t.id, stage="group", group_id=g.id, team_a_id=ta.id, team_b_id=tb.id,
                                status="completed", score_a=20, score_b=10, points_a=3, points_b=0))
    db_session.add(models.Match(tournament_id=t.id, stage="group", group_id=g.id, team_a_id=tb.id, team_b_id=tc.id,
                                status="completed", score_a=15, score_b=15, points_a=1, points_b=1, is_draw=True))
    db_session.add(models.Match(tournament_id=t.id, stage="group", group_id=g.id, team_a_id=ta.id, team_b_id=tc.id,
                                status="completed", score_a=12, score_b=8, points_a=3, points_b=0))
    db_session.commit()

    rows = standings.compute_group_standings(db_session, t, g)
    assert rows[0]["team_name"] == "A"
    assert rows[0]["points"] == 6


def test_standings_export_separates_groups_into_sheets(db_session, admin_client):
    t = models.Tournament(name="Cup", status="active")
    db_session.add(t)
    db_session.commit()
    db_session.refresh(t)

    teams = [make_team(db_session, name) for name in ("A1", "A2", "B1", "B2")]
    for team in teams:
        db_session.add(models.TournamentTeam(tournament_id=t.id, team_id=team.id))
    g1 = models.Group(tournament_id=t.id, name="Group A")
    g2 = models.Group(tournament_id=t.id, name="Group B")
    db_session.add_all([g1, g2])
    db_session.commit()
    db_session.refresh(g1)
    db_session.refresh(g2)
    for team in teams[:2]:
        db_session.add(models.GroupTeam(group_id=g1.id, team_id=team.id))
    for team in teams[2:]:
        db_session.add(models.GroupTeam(group_id=g2.id, team_id=team.id))
    db_session.add(models.Match(
        tournament_id=t.id, stage="group", group_id=g1.id,
        team_a_id=teams[0].id, team_b_id=teams[1].id,
        status="completed", score_a=7, score_b=3,
    ))
    db_session.add(models.Match(
        tournament_id=t.id, stage="group", group_id=g2.id,
        team_a_id=teams[2].id, team_b_id=teams[3].id,
        status="completed", score_a=4, score_b=4, is_draw=True,
    ))
    db_session.commit()

    r = admin_client.get(f"/api/reports/standings/{t.id}.xlsx")

    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    try:
        assert wb.sheetnames == ["Group A", "Group B"]
        group_a_rows = list(wb["Group A"].iter_rows(values_only=True))
        group_b_rows = list(wb["Group B"].iter_rows(values_only=True))
        assert group_a_rows[1][1] == "A1"
        assert group_a_rows[1][6] == 3
        assert group_b_rows[1][6] == 1
    finally:
        wb.close()


def test_knockout_generation_with_byes_and_advance(db_session):
    t = models.Tournament(name="KO", status="active")
    db_session.add(t)
    db_session.commit()
    db_session.refresh(t)

    teams = [make_team(db_session, f"T{i}") for i in range(1, 7)]
    qids = [x.id for x in teams]
    brackets.generate_bracket(db_session, t, qids)

    rounds = db_session.query(models.BracketRound).filter_by(tournament_id=t.id).all()
    assert len(rounds) >= 3
    matches = db_session.query(models.Match).filter_by(tournament_id=t.id, stage="knockout").all()
    assert len(matches) >= 2

    m = matches[0]
    m.score_a = 10
    m.score_b = 5
    match_service.complete_match(db_session, m, host_id=None)
    db_session.refresh(m)
    assert m.winner_team_id is not None


def test_knockout_no_draw_requires_winner(db_session):
    t = models.Tournament(name="KO2", status="active")
    db_session.add(t)
    db_session.commit()
    db_session.refresh(t)

    a = make_team(db_session, "A")
    b = make_team(db_session, "B")
    m = models.Match(tournament_id=t.id, stage="knockout", team_a_id=a.id, team_b_id=b.id,
                     status="in_progress", score_a=10, score_b=10)
    db_session.add(m)
    db_session.commit()

    from app.game import GameError
    try:
        match_service.complete_match(db_session, m, host_id=None)
        assert False, "expected GameError"
    except GameError as e:
        assert e.key == "knockout_no_draw"


def test_localization_default_language_persistence(db_session):
    assert i18n.get_default_language(db_session) == "ar-EG"
    i18n.set_default_language(db_session, "en")
    assert i18n.get_default_language(db_session) == "en"
    tr_ar = i18n.Translator("ar-EG")
    tr_en = i18n.Translator("en")
    assert tr_ar.dir == "rtl"
    assert tr_en.dir == "ltr"
    assert tr_ar.t("login") != tr_en.t("login")
